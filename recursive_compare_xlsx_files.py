# -*- coding: utf-8 -*-
# python-version: 3.6.2
#
# recursive_compare_xlsx_files.py
#
#   usage: python recursive_compare_xlsx_files.py <original xlsx files root dir> <target xlsx files root dir>
#
import os
import sys
import re
from openpyxl import load_workbook

def print_flush(v):
    print(v)
    sys.stdout.flush()


def compare_xlsx_files(xlsx_file1, xlsx_file2):
    if xlsx_file1 is None or xlsx_file2 is None:
        raise RuntimeError('SKIP: %s, %s' % (xlsx_file1, xlsx_file2))

    wb1 = load_workbook(xlsx_file1, read_only=True)
    wb2 = load_workbook(xlsx_file2, read_only=True)

    for sheet_name in wb1.get_sheet_names():
        if sheet_name in wb2:
            ws1 = wb1[sheet_name]
            ws2 = wb2[sheet_name]
            ws1_rows = tuple(ws1.rows)
            ws2_rows = tuple(ws2.rows)
            range_max_row = range(1,ws1.max_row)
            range_max_column = range(1,ws1.max_column)

            for i in range_max_row:
                ws1_row = ws1_rows[i]
                ws2_row = ws2_rows[i]

                for j in range_max_column:
                    v1 = ws1_row[j].value
                    v2 = ws2_row[j].value

                    if v1 != v2:
                        raise RuntimeError('ERROR: %s:{sheet_name=\'%s\', cell(\'%s\')}=%s, %s:{sheet_name=\'%s\', cell(\'%s\')}=%s unmatched' %\
                            (xlsx_file1, sheet_name, ws1_row[j].coordinate, v1, xlsx_file2, sheet_name, ws2_row[j].coordinate, v2))
        else:
            raise RuntimeError('ERROR: sheet_name %s not found at %s ' % (sheet_name, xlsx_file2))

    print_flush('DONE: %s, %s' % (xlsx_file1, xlsx_file2))


def find_path(base_filename, target_dir2):
    for root, dirs, files in os.walk(target_dir2):
        for filename in files:
            if (base_filename == filename):
                filepath = os.path.join(root,filename)
                return filepath

def find_and_compare(filepath, target_dir2):
    reg_xlsx_file = re.compile('.*\.xlsx$')

    if (not os.path.isdir(filepath)):
        is_xlsx_file = reg_xlsx_file.match(filepath)

        if is_xlsx_file:
            base_filename = os.path.basename(filepath)
            file1 = filepath
            file2 = find_path(base_filename, target_dir2)
            try:
                compare_xlsx_files(file1, file2)
            except Exception as error:
                print_flush(error)

def main():
    original_dir = sys.argv[1]
    target_dir = sys.argv[2]
    for root, dirs, files in os.walk(original_dir):
        for filename in files:
            filepath = os.path.join(root,filename)
            find_and_compare(filepath, target_dir)

if __name__ == '__main__':
    main()
